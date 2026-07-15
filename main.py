import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import threading
from datetime import datetime

from database import Database
from gmail_service import GmailService
from templates import get_stock_email_template, get_stock_email_template_table
from config import EMAIL_SETTINGS, EMAIL_TEMPLATE

class StockEmailSenderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gmail Stock Email Sender")
        self.root.geometry("920x780")
        self.root.configure(bg="#f5f5f5")
        
        self.db = Database()
        self.gmail_service = GmailService()
        self.stock_data = []
        self.sending = False
        self.attachment_path = None
        
        self._create_ui()
        self._refresh_accounts_list()
        self._refresh_customer_count()
        self._refresh_daily_stats()
        self._refresh_gmail_dropdown()
    
    def _create_ui(self):
        style = ttk.Style()
        style.configure('Notebook.TNotebook', background='#f5f5f5')
        style.configure('TNotebook.Tab', padding=[12, 8], font=('Segoe UI', 10))
        style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'), background='#f5f5f5')
        style.configure('TButton', padding=[10, 5], font=('Segoe UI', 10))
        style.configure('Stats.TLabel', font=('Segoe UI', 14, 'bold'), foreground='#1e3c72')
        
        stats_frame = ttk.Frame(self.root, padding=10)
        stats_frame.pack(fill='x', padx=10)
        
        ttk.Label(stats_frame, text="📊 Daily Email Stats:", font=('Segoe UI', 12, 'bold')).pack(side='left', padx=5)
        
        self.daily_sent_label = ttk.Label(stats_frame, text="Sent Today: 0", style='Stats.TLabel')
        self.daily_sent_label.pack(side='left', padx=20)
        
        self.daily_limit_label = ttk.Label(stats_frame, text="Daily Limit: 0", font=('Segoe UI', 12), foreground='#666')
        self.daily_limit_label.pack(side='left', padx=20)
        
        self.daily_remaining_label = ttk.Label(stats_frame, text="Remaining: 0", font=('Segoe UI', 12), foreground='#28a745')
        self.daily_remaining_label.pack(side='left', padx=20)
        
        ttk.Button(stats_frame, text="🔄 Refresh", command=self._refresh_daily_stats).pack(side='right', padx=5)
        
        ttk.Separator(self.root, orient='horizontal').pack(fill='x', padx=10)
        
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        accounts_frame = ttk.Frame(notebook)
        notebook.add(accounts_frame, text='📧 Gmail Accounts')
        self._create_accounts_tab(accounts_frame)
        
        customers_frame = ttk.Frame(notebook)
        notebook.add(customers_frame, text='👥 Customers')
        self._create_customers_tab(customers_frame)
        
        stocks_frame = ttk.Frame(notebook)
        notebook.add(stocks_frame, text='📈 Stock Data')
        self._create_stocks_tab(stocks_frame)
        
        send_frame = ttk.Frame(notebook)
        notebook.add(send_frame, text='📨 Send Emails')
        self._create_send_tab(send_frame)
    
    def _create_accounts_tab(self, parent):
        padding = {'padx': 15, 'pady': 10}
        
        header = ttk.Label(parent, text="Manage Gmail Accounts", style='Header.TLabel')
        header.grid(row=0, column=0, columnspan=2, **padding, sticky='w')
        
        ttk.Label(parent, text="Note: Use App Password (not regular password)", 
                  font=('Segoe UI', 9, 'italic'), foreground='#666').grid(
            row=1, column=0, columnspan=2, padx=15, sticky='w')
        
        form_frame = ttk.LabelFrame(parent, text="Add New Account", padding=10)
        form_frame.grid(row=2, column=0, columnspan=2, padx=15, pady=10, sticky='ew')
        
        ttk.Label(form_frame, text="Gmail Address:").grid(row=0, column=0, sticky='w', pady=5)
        self.email_entry = ttk.Entry(form_frame, width=40)
        self.email_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="App Password:").grid(row=1, column=0, sticky='w', pady=5)
        self.password_entry = ttk.Entry(form_frame, width=40, show="*")
        self.password_entry.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Button(form_frame, text="Add Account", command=self._add_account).grid(
            row=2, column=0, columnspan=2, pady=10)
        
        list_frame = ttk.LabelFrame(parent, text="Active Accounts", padding=10)
        list_frame.grid(row=3, column=0, columnspan=2, padx=15, pady=10, sticky='nsew')
        
        self.accounts_tree = ttk.Treeview(list_frame, columns=('email', 'sent_today', 'limit'), 
                                          show='headings', height=6)
        self.accounts_tree.heading('email', text='Email')
        self.accounts_tree.heading('sent_today', text='Sent Today')
        self.accounts_tree.heading('limit', text='Daily Limit')
        self.accounts_tree.column('email', width=250)
        self.accounts_tree.column('sent_today', width=100)
        self.accounts_tree.column('limit', width=100)
        self.accounts_tree.pack(fill='both', expand=True)
        
        btn_frame = ttk.Frame(list_frame)
        btn_frame.pack(fill='x', pady=10)
        ttk.Button(btn_frame, text="Delete Selected", command=self._delete_account).pack(side='left')
        ttk.Button(btn_frame, text="Refresh", command=self._refresh_accounts_list).pack(side='left', padx=5)
        
        help_frame = ttk.LabelFrame(parent, text="How to Create App Password", padding=10)
        help_frame.grid(row=4, column=0, columnspan=2, padx=15, pady=10, sticky='ew')
        
        help_text = """
1. Go to https://myaccount.google.com/
2. Enable 2-Step Verification (if not already done)
3. Search for "App Passwords" or go to Security > App Passwords
4. Select "Mail" and your device
5. Generate and copy the 16-character password
6. Paste it here (not your regular Gmail password)
        """
        ttk.Label(help_frame, text=help_text, justify='left', font=('Segoe UI', 9)).pack(anchor='w')
    
    def _create_customers_tab(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)
        
        padding = {'padx': 15, 'pady': 10}
        
        header = ttk.Label(parent, text="Customer Management", style='Header.TLabel')
        header.grid(row=0, column=0, **padding, sticky='w')
        
        top_frame = ttk.Frame(parent)
        top_frame.grid(row=1, column=0, padx=15, pady=5, sticky='ew')
        top_frame.columnconfigure(0, weight=1)
        top_frame.columnconfigure(1, weight=1)
        top_frame.columnconfigure(2, weight=1)
        
        upload_frame = ttk.LabelFrame(top_frame, text="Upload CSV", padding=10)
        upload_frame.grid(row=0, column=0, padx=5, sticky='nsew')
        
        ttk.Label(upload_frame, text="Format: store_code, email, mobile_number", 
                  font=('Segoe UI', 8, 'italic'), foreground='#666').pack(anchor='w')
        ttk.Button(upload_frame, text="Upload CSV/Excel", command=self._upload_customers_csv).pack(pady=5, fill='x')
        
        stats_frame = ttk.LabelFrame(top_frame, text="Statistics", padding=10)
        stats_frame.grid(row=0, column=1, padx=5, sticky='nsew')
        
        self.customer_count_label = ttk.Label(stats_frame, text="Total: 0", font=('Segoe UI', 10))
        self.customer_count_label.pack(anchor='w')
        
        self.selected_count_label = ttk.Label(stats_frame, text="Selected: 0", font=('Segoe UI', 10))
        self.selected_count_label.pack(anchor='w')
        
        self.unsent_count_label = ttk.Label(stats_frame, text="Unsent: 0", font=('Segoe UI', 10))
        self.unsent_count_label.pack(anchor='w')
        
        add_frame = ttk.LabelFrame(top_frame, text="Add Customer", padding=10)
        add_frame.grid(row=0, column=2, padx=5, sticky='nsew')
        
        ttk.Label(add_frame, text="Email:").grid(row=0, column=0, sticky='w')
        self.manual_email_entry = ttk.Entry(add_frame, width=20)
        self.manual_email_entry.grid(row=0, column=1, padx=2)
        
        ttk.Label(add_frame, text="Store Code:").grid(row=1, column=0, sticky='w')
        self.manual_store_code_entry = ttk.Entry(add_frame, width=20)
        self.manual_store_code_entry.grid(row=1, column=1, padx=2)
        
        ttk.Label(add_frame, text="Mobile:").grid(row=2, column=0, sticky='w')
        self.manual_mobile_entry = ttk.Entry(add_frame, width=20)
        self.manual_mobile_entry.grid(row=2, column=1, padx=2)
        
        ttk.Button(add_frame, text="Add", command=self._add_customer_manual).grid(row=3, column=0, columnspan=2, pady=5)
        
        list_frame = ttk.LabelFrame(parent, text="Customer List", padding=5)
        list_frame.grid(row=2, column=0, padx=15, pady=10, sticky='nsew')
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(1, weight=1)
        
        list_toolbar = ttk.Frame(list_frame)
        list_toolbar.grid(row=0, column=0, sticky='ew', pady=5)
        
        ttk.Button(list_toolbar, text="Select All", command=self._select_all_customers).pack(side='left', padx=2)
        ttk.Button(list_toolbar, text="Deselect All", command=self._deselect_all_customers).pack(side='left', padx=2)
        ttk.Button(list_toolbar, text="Delete Selected", command=self._delete_selected_customers).pack(side='left', padx=2)
        ttk.Button(list_toolbar, text="Mark Selected Unsce", command=self._mark_selected_unsent).pack(side='left', padx=2)
        ttk.Button(list_toolbar, text="Refresh List", command=self._refresh_customers_list).pack(side='left', padx=2)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._filter_customers)
        ttk.Label(list_toolbar, text="Search:").pack(side='left', padx=(10, 2))
        search_entry = ttk.Entry(list_toolbar, textvariable=self.search_var, width=20)
        search_entry.pack(side='left')
        
        columns = ('select', 'store_code', 'email', 'mobile_number', 'sent_history', 'status')
        self.customers_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        self.customers_tree.heading('select', text='Select')
        self.customers_tree.heading('store_code', text='Store Code')
        self.customers_tree.heading('email', text='Email')
        self.customers_tree.heading('mobile_number', text='Mobile Number')
        self.customers_tree.heading('sent_history', text='Sent (Last 3 Days)')
        self.customers_tree.heading('status', text='Status')
        
        self.customers_tree.column('select', width=50, anchor='center')
        self.customers_tree.column('store_code', width=80, anchor='center')
        self.customers_tree.column('email', width=180)
        self.customers_tree.column('mobile_number', width=100, anchor='center')
        self.customers_tree.column('sent_history', width=150, anchor='center')
        self.customers_tree.column('status', width=80, anchor='center')
        
        self.customers_tree.grid(row=1, column=0, sticky='nsew')
        
        self.customers_tree.bind('<Button-1>', self._toggle_customer_selection)
        
        vsb = ttk.Scrollbar(list_frame, orient='vertical', command=self.customers_tree.yview)
        vsb.grid(row=1, column=1, sticky='ns')
        self.customers_tree.configure(yscrollcommand=vsb.set)
        
        bottom_frame = ttk.Frame(parent)
        bottom_frame.grid(row=3, column=0, padx=15, pady=10, sticky='ew')
        
        filter_frame = ttk.LabelFrame(bottom_frame, text="Filter & Clean", padding=5)
        filter_frame.pack(side='left', padx=5)
        
        ttk.Button(filter_frame, text="Select Invalid Emails", command=self._select_invalid_emails).pack(side='left', padx=2)
        ttk.Button(filter_frame, text="Select Failed 2+ Times", command=self._select_failed_emails).pack(side='left', padx=2)
        ttk.Button(filter_frame, text="Delete Selected", command=self._delete_selected_customers).pack(side='left', padx=2)
        
        export_frame = ttk.Frame(bottom_frame)
        export_frame.pack(side='left', padx=20)
        
        ttk.Button(export_frame, text="Export Selected", command=self._export_selected_emails).pack(side='left', padx=2)
        ttk.Button(export_frame, text="Export All", command=self._export_all_emails).pack(side='left', padx=2)
        
        ttk.Button(bottom_frame, text="Clear All Customers", command=self._clear_customers).pack(side='right', padx=5)
        ttk.Button(bottom_frame, text="Reset Sent Status", command=self._reset_sent_status).pack(side='right', padx=5)
    
    def _create_stocks_tab(self, parent):
        padding = {'padx': 15, 'pady': 10}
        
        header = ttk.Label(parent, text="Stock File Attachment", style='Header.TLabel')
        header.grid(row=0, column=0, **padding, sticky='w')
        
        upload_frame = ttk.LabelFrame(parent, text="Select Stock File", padding=20)
        upload_frame.grid(row=1, column=0, padx=15, pady=10, sticky='ew')
        
        ttk.Label(upload_frame, text="Select your stock file (Excel or CSV) to attach to emails:", 
                  font=('Segoe UI', 10)).pack(anchor='w', pady=5)
        
        ttk.Label(upload_frame, text="Supported formats: .xlsx, .xls, .csv", 
                  font=('Segoe UI', 9, 'italic'), foreground='#666').pack(anchor='w', pady=5)
        
        self.stock_file_label = ttk.Label(upload_frame, text="No file selected", 
                                          font=('Segoe UI', 11), foreground='#666')
        self.stock_file_label.pack(anchor='w', pady=10)
        
        ttk.Button(upload_frame, text="📁 Select Stock File", 
                   command=self._upload_stocks_csv).pack(pady=10)
        
        ttk.Button(upload_frame, text="🗑 Clear Selection", 
                   command=self._clear_stocks).pack(pady=5)
        
        info_frame = ttk.LabelFrame(parent, text="Instructions", padding=15)
        info_frame.grid(row=2, column=0, padx=15, pady=10, sticky='ew')
        
        instructions = """
1. Click "Select Stock File" above
2. Choose your Excel (.xlsx/.xls) or CSV file
3. The file will be attached to all emails
4. Go to "Send Emails" tab to send

Note: The stock file will be sent as an attachment.
The email body contains a standard message about new stock arrival.
        """
        ttk.Label(info_frame, text=instructions, justify='left', 
                  font=('Segoe UI', 10)).pack(anchor='w')
    
    def _create_send_tab(self, parent):
        padding = {'padx': 15, 'pady': 10}
        
        header = ttk.Label(parent, text="Send Emails", style='Header.TLabel')
        header.grid(row=0, column=0, **padding, sticky='w')
        
        settings_frame = ttk.LabelFrame(parent, text="Email Settings", padding=10)
        settings_frame.grid(row=1, column=0, padx=15, pady=5, sticky='ew')
        
        ttk.Label(settings_frame, text="Email Subject:").grid(row=0, column=0, sticky='w', pady=2)
        self.subject_entry = ttk.Entry(settings_frame, width=50)
        self.subject_entry.insert(0, EMAIL_TEMPLATE['subject'])
        self.subject_entry.grid(row=0, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(settings_frame, text="Short Message:").grid(row=1, column=0, sticky='w', pady=2)
        self.custom_message_entry = ttk.Entry(settings_frame, width=50)
        self.custom_message_entry.insert(0, "Many new and short items included in the list and marked.")
        self.custom_message_entry.grid(row=1, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(settings_frame, text="Batch Size:").grid(row=2, column=0, sticky='w', pady=2)
        self.batch_size_entry = ttk.Spinbox(settings_frame, from_=10, to=100, width=10)
        self.batch_size_entry.set(EMAIL_SETTINGS['batch_size'])
        self.batch_size_entry.grid(row=2, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(settings_frame, text="Delay Between Emails (sec):").grid(row=3, column=0, sticky='w', pady=2)
        self.delay_entry = ttk.Spinbox(settings_frame, from_=1, to=30, width=10)
        self.delay_entry.set(EMAIL_SETTINGS['delay_between_emails'])
        self.delay_entry.grid(row=3, column=1, padx=5, pady=2, sticky='w')
        
        ttk.Label(settings_frame, text="Delay Between Batches (sec):").grid(row=4, column=0, sticky='w', pady=2)
        self.batch_delay_entry = ttk.Spinbox(settings_frame, from_=30, to=300, width=10)
        self.batch_delay_entry.set(EMAIL_SETTINGS['delay_between_batches'])
        self.batch_delay_entry.grid(row=4, column=1, padx=5, pady=2, sticky='w')
        
        send_frame = ttk.LabelFrame(parent, text="🚀 SEND EMAILS", padding=15)
        send_frame.grid(row=2, column=0, padx=15, pady=10, sticky='ew')
        
        gmail_select_frame = ttk.Frame(send_frame)
        gmail_select_frame.pack(fill='x', pady=5)
        
        ttk.Label(gmail_select_frame, text="Select Gmail Account:").pack(side='left', padx=5)
        self.gmail_combo = ttk.Combobox(gmail_select_frame, width=35, state='readonly')
        self.gmail_combo.pack(side='left', padx=5)
        self.gmail_combo.bind('<<ComboboxSelected>>', self._on_gmail_selected)
        
        self.gmail_accounts_list = []
        
        self.send_btn = ttk.Button(send_frame, text="SEND NOW", command=self._start_sending, width=20)
        self.send_btn.pack(side='left', padx=10, pady=5)
        
        self.stop_btn = ttk.Button(send_frame, text="STOP", command=self._stop_sending, state='disabled', width=10)
        self.stop_btn.pack(side='left', padx=5, pady=5)
        
        ttk.Button(send_frame, text="Refresh", command=self._refresh_customer_count, width=10).pack(side='left', padx=5, pady=5)
        
        attachment_frame = ttk.LabelFrame(parent, text="Stock File Attachment", padding=10)
        attachment_frame.grid(row=3, column=0, padx=15, pady=5, sticky='ew')
        
        self.attachment_path_label = ttk.Label(attachment_frame, text="No file selected", foreground='#666')
        self.attachment_path_label.pack(side='left', padx=5)
        
        ttk.Button(attachment_frame, text="Select File", command=self._select_attachment).pack(side='left', padx=5)
        ttk.Button(attachment_frame, text="Clear", command=self._clear_attachment).pack(side='left', padx=5)
        
        target_frame = ttk.LabelFrame(parent, text="Email Recipients", padding=10)
        target_frame.grid(row=4, column=0, padx=15, pady=5, sticky='ew')
        
        self.send_to_selected_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(target_frame, text="Send to selected customers only", 
                        variable=self.send_to_selected_var).pack(anchor='w')
        
        progress_frame = ttk.LabelFrame(parent, text="Progress", padding=10)
        progress_frame.grid(row=5, column=0, padx=15, pady=5, sticky='ew')
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate', length=400)
        self.progress_bar.pack(fill='x', pady=5)
        
        self.progress_label = ttk.Label(progress_frame, text="Ready to send")
        self.progress_label.pack()
        
        self.send_stats_label = ttk.Label(progress_frame, text="Sent: 0 | Failed: 0 | Remaining: 0")
        self.send_stats_label.pack()
        
        preview_frame = ttk.LabelFrame(parent, text="Email Preview", padding=5)
        preview_frame.grid(row=6, column=0, padx=15, pady=5, sticky='ew')
        
        self.preview_text = scrolledtext.ScrolledText(preview_frame, height=4, wrap='word')
        self.preview_text.pack(fill='x')
        
        ttk.Button(preview_frame, text="Preview", command=self._preview_email).pack(pady=2)
    
    def _refresh_daily_stats(self):
        accounts = self.db.get_gmail_accounts()
        total_sent_today = 0
        active_count = 0
        
        for acc in accounts:
            acc_id, email, password, is_active, sent_today, last_reset = acc
            if is_active:
                active_count += 1
                total_sent_today += sent_today
        
        total_limit = active_count * EMAIL_SETTINGS['max_emails_per_account_per_day']
        remaining = total_limit - total_sent_today
        
        self.daily_sent_label.config(text=f"Sent Today: {total_sent_today}")
        self.daily_limit_label.config(text=f"Daily Limit: {total_limit}")
        self.daily_remaining_label.config(text=f"Remaining: {remaining}")
    
    def _refresh_gmail_dropdown(self):
        accounts = self.db.get_all_gmail_accounts_for_dropdown()
        self.gmail_accounts_list = accounts
        account_names = [f"{acc[1]} (Sent: {self.db.get_account_sent_count(acc[0])})" for acc in accounts]
        self.gmail_combo['values'] = account_names
        if account_names:
            self.gmail_combo.current(0)
    
    def _on_gmail_selected(self, event):
        pass
        
        if remaining < 50:
            self.daily_remaining_label.config(foreground='#dc3545')
        elif remaining < 200:
            self.daily_remaining_label.config(foreground='#ffc107')
        else:
            self.daily_remaining_label.config(foreground='#28a745')
    
    def _add_account(self):
        email = self.email_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not email or not password:
            messagebox.showerror("Error", "Please enter both email and password")
            return
        
        if '@' not in email:
            messagebox.showerror("Error", "Please enter a valid email address")
            return
        
        if self.db.add_gmail_account(email, password):
            messagebox.showinfo("Success", f"Account {email} added successfully!")
            self.email_entry.delete(0, tk.END)
            self.password_entry.delete(0, tk.END)
            self._refresh_accounts_list()
            self._refresh_gmail_dropdown()
        else:
            messagebox.showerror("Error", "Account already exists or failed to add")
    
    def _toggle_custom_body(self):
        if self.use_custom_body_var.get():
            self.custom_body_text.config(state='normal')
        else:
            self.custom_body_text.config(state='disabled')
    
    def _select_attachment(self):
        file_path = filedialog.askopenfilename(
            title="Select Attachment File",
            filetypes=[
                ("All Files", "*.*"),
                ("CSV Files", "*.csv"),
                ("Excel Files", "*.xlsx *.xls"),
                ("PDF Files", "*.pdf"),
                ("Text Files", "*.txt")
            ]
        )
        
        if file_path:
            self.attachment_path = file_path
            filename = file_path.split('/')[-1].split('\\')[-1]
            self.attachment_path_label.config(text=filename, foreground='#28a745')
    
    def _clear_attachment(self):
        self.attachment_path = None
        self.attachment_path_label.config(text="No file selected", foreground='#666')
    
    def _delete_account(self):
        selected = self.accounts_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an account to delete")
            return
        
        item = self.accounts_tree.item(selected[0])
        email = item['values'][0]
        
        if messagebox.askyesno("Confirm", f"Delete account {email}?"):
            self.db.delete_gmail_account(email)
            self._refresh_accounts_list()
            self._refresh_gmail_dropdown()
    
    def _refresh_accounts_list(self):
        for item in self.accounts_tree.get_children():
            self.accounts_tree.delete(item)
        
        accounts = self.db.get_gmail_accounts()
        for acc in accounts:
            acc_id, email, password, is_active, sent_today, last_reset = acc
            if last_reset < datetime.now().date().isoformat():
                sent_today = 0
            self.accounts_tree.insert('', 'end', values=(
                email,
                sent_today,
                EMAIL_SETTINGS['max_emails_per_account_per_day']
            ))
    
    def _upload_customers_csv(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV/Excel Files", "*.csv *.xlsx *.xls"), ("CSV Files", "*.csv"), ("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return
        
        customers = []
        
        try:
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        store_code = row.get('store_code', '') or row.get('Store_Code', '') or row.get('StoreCode', '') or row.get('Store Code', '')
                        email = row.get('email', '') or row.get('Email', '') or row.get('EMAIL', '')
                        mobile_number = row.get('mobile_number', '') or row.get('Mobile_Number', '') or row.get('MobileNumber', '') or row.get('Mobile Number', '') or row.get('mobile', '') or row.get('Mobile', '')
                        customers.append({
                            'email': email.strip(),
                            'store_code': store_code.strip() if store_code else None,
                            'mobile_number': mobile_number.strip() if mobile_number else None
                        })
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror("Error", "Excel support requires openpyxl library.\n\nInstall with: pip install openpyxl")
                    return
                
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip().replace('_', '').replace(' ', '') if cell.value else '')
                
                store_code_col = None
                email_col = None
                mobile_col = None
                
                for idx, header in enumerate(headers):
                    if header in ['email', 'emails', 'e-mail']:
                        email_col = idx
                    elif header in ['storecode', 'store_code', 'store']:
                        store_code_col = idx
                    elif header in ['mobilenumber', 'mobile_number', 'mobile', 'phone']:
                        mobile_col = idx
                
                if email_col is None and store_code_col is None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            customers.append({
                                'store_code': str(row[0]).strip() if row[0] else None,
                                'email': str(row[1]).strip() if len(row) > 1 and row[1] else None,
                                'mobile_number': str(row[2]).strip() if len(row) > 2 and row[2] else None
                            })
                else:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row:
                            email = str(row[email_col]).strip() if email_col is not None and len(row) > email_col and row[email_col] else None
                            store_code = str(row[store_code_col]).strip() if store_code_col is not None and len(row) > store_code_col and row[store_code_col] else None
                            mobile_number = str(row[mobile_col]).strip() if mobile_col is not None and len(row) > mobile_col and row[mobile_col] else None
                            if email:
                                customers.append({
                                    'email': email,
                                    'store_code': store_code,
                                    'mobile_number': mobile_number
                                })
                
                wb.close()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read file: {str(e)}")
            return
        
        valid_customers = [c for c in customers if c['email']]
        added = self.db.add_customers_batch(valid_customers)
        
        messagebox.showinfo("Success", f"Added {added} customers\n({len(customers) - added} duplicates skipped)")
        self._refresh_customers_list()
    
    def _add_customer_manual(self):
        email = self.manual_email_entry.get().strip()
        store_code = self.manual_store_code_entry.get().strip()
        mobile_number = self.manual_mobile_entry.get().strip()
        
        if not email:
            messagebox.showerror("Error", "Email is required")
            return
        
        if self.db.add_customer(email, store_code, mobile_number):
            self.manual_email_entry.delete(0, tk.END)
            self.manual_store_code_entry.delete(0, tk.END)
            self.manual_mobile_entry.delete(0, tk.END)
            self._refresh_customers_list()
        else:
            messagebox.showerror("Error", "Customer already exists")
    
    def _refresh_customers_list(self, filter_text=None):
        for item in self.customers_tree.get_children():
            self.customers_tree.delete(item)
        
        conn = self.db._get_connection()
        cursor = conn.cursor()
        
        fail_counts = {}
        cursor.execute('''
            SELECT customer_id, COUNT(*) as cnt 
            FROM email_logs 
            WHERE status = 'failed'
            GROUP BY customer_id
        ''')
        for row in cursor.fetchall():
            fail_counts[row[0]] = row[1]
        
        query = '''
            SELECT c.id, c.email, c.store_code, c.mobile_number, 
                   CASE WHEN el_sent.id IS NOT NULL THEN 'Sent' ELSE 'Unsent' END as status
            FROM customers c
            LEFT JOIN email_logs el_sent ON c.id = el_sent.customer_id AND el_sent.status = 'sent'
        '''
        
        if filter_text:
            query += " WHERE c.email LIKE ? OR c.store_code LIKE ? OR c.mobile_number LIKE ?"
            cursor.execute(query, (f'%{filter_text}%', f'%{filter_text}%', f'%{filter_text}%'))
        else:
            cursor.execute(query)
        
        customers = cursor.fetchall()
        conn.close()
        
        for cust in customers:
            cust_id, email, store_code, mobile_number, status = cust
            
            sent_dates = self.db.get_sent_dates_for_customer(cust_id, days=3)
            sent_history = ', '.join(sent_dates) if sent_dates else '-'
            
            fail_count = fail_counts.get(cust_id, 0)
            if status == 'Sent':
                status_display = 'Sent'
            elif fail_count >= 2:
                status_display = f'Failed x{fail_count}'
            elif fail_count > 0:
                status_display = f'Failed x{fail_count}'
            else:
                status_display = 'Unsent'
            
            self.customers_tree.insert('', 'end', iid=cust_id, values=(
                '☐',
                store_code or '',
                email,
                mobile_number or '',
                sent_history,
                status_display
            ))
        
        total = self.db.get_customer_count()
        unsent = self.db.get_unsent_count()
        self.customer_count_label.config(text=f"Total: {total}")
        self.unsent_count_label.config(text=f"Unsent: {unsent}")
        self._update_selected_count()
        
        if hasattr(self, 'send_stats_label'):
            self.send_stats_label.config(text=f"Sent: 0 | Failed: 0 | Remaining: {unsent}")
    
    def _filter_customers(self, *args):
        filter_text = self.search_var.get().strip()
        self._refresh_customers_list(filter_text if filter_text else None)
    
    def _toggle_customer_selection(self, event):
        region = self.customers_tree.identify('region', event.x, event.y)
        if region == 'cell':
            column = self.customers_tree.identify_column(event.x)
            if column == '#1':
                item = self.customers_tree.identify_row(event.y)
                if item:
                    current = self.customers_tree.item(item)['values'][0]
                    new_val = '☑' if current == '☐' else '☐'
                    self.customers_tree.item(item, values=(new_val,) + tuple(self.customers_tree.item(item)['values'][1:]))
                    self._update_selected_count()
    
    def _update_selected_count(self):
        selected = sum(1 for item in self.customers_tree.get_children() 
                      if self.customers_tree.item(item)['values'][0] == '☑')
        self.selected_count_label.config(text=f"Selected: {selected}")
    
    def _select_all_customers(self):
        for item in self.customers_tree.get_children():
            values = list(self.customers_tree.item(item)['values'])
            values[0] = '☑'
            self.customers_tree.item(item, values=values)
        self._update_selected_count()
    
    def _deselect_all_customers(self):
        for item in self.customers_tree.get_children():
            values = list(self.customers_tree.item(item)['values'])
            values[0] = '☐'
            self.customers_tree.item(item, values=values)
        self._update_selected_count()
    
    def _get_selected_customer_ids(self):
        return [item for item in self.customers_tree.get_children() 
                if self.customers_tree.item(item)['values'][0] == '☑']
    
    def _delete_selected_customers(self):
        selected = self._get_selected_customer_ids()
        if not selected:
            messagebox.showwarning("Warning", "No customers selected")
            return
        
        if messagebox.askyesno("Confirm", f"Delete {len(selected)} selected customers?"):
            conn = self.db._get_connection()
            cursor = conn.cursor()
            for cust_id in selected:
                cursor.execute('DELETE FROM customers WHERE id = ?', (cust_id,))
                cursor.execute('DELETE FROM email_logs WHERE customer_id = ?', (cust_id,))
            conn.commit()
            conn.close()
            self._refresh_customers_list()
    
    def _mark_selected_unsent(self):
        selected = self._get_selected_customer_ids()
        if not selected:
            messagebox.showwarning("Warning", "No customers selected")
            return
        
        conn = self.db._get_connection()
        cursor = conn.cursor()
        for cust_id in selected:
            cursor.execute('DELETE FROM email_logs WHERE customer_id = ?', (cust_id,))
        conn.commit()
        conn.close()
        self._refresh_customers_list()
        messagebox.showinfo("Success", f"Marked {len(selected)} customers as unsent")
    
    def _select_invalid_emails(self):
        self._deselect_all_customers()
        
        invalid_emails = self.db.get_invalid_emails()
        
        count = 0
        for item in self.customers_tree.get_children():
            values = self.customers_tree.item(item)['values']
            if values[1] in invalid_emails:
                values = list(values)
                values[0] = '☑'
                self.customers_tree.item(item, values=values)
                count += 1
        
        self._update_selected_count()
        if count > 0:
            messagebox.showinfo("Selected", f"Selected {count} invalid email addresses")
        else:
            messagebox.showinfo("No Invalid Emails", "No invalid email addresses found")
    
    def _select_failed_emails(self):
        self._deselect_all_customers()
        
        failed_customers = self.db.get_customers_with_failures(min_failures=2)
        failed_ids = [row[0] for row in failed_customers]
        
        count = 0
        for item in self.customers_tree.get_children():
            if item in failed_ids:
                values = list(self.customers_tree.item(item)['values'])
                values[0] = '☑'
                self.customers_tree.item(item, values=values)
                count += 1
        
        self._update_selected_count()
        if count > 0:
            messagebox.showinfo("Selected", f"Selected {count} emails that failed 2+ times")
        else:
            messagebox.showinfo("No Failed Emails", "No emails with 2+ failures found")
    
    def _export_selected_emails(self):
        selected = self._get_selected_customer_ids()
        if not selected:
            messagebox.showwarning("Warning", "No customers selected")
            return
        
        emails = [self.customers_tree.item(item)['values'][1] for item in selected]
        
        file_path = filedialog.asksaveasfilename(
            defaultextension='.txt',
            filetypes=[("Text Files", "*.txt"), ("CSV Files", "*.csv")],
            title="Export Selected Emails"
        )
        
        if file_path:
            with open(file_path, 'w') as f:
                f.write('\n'.join(emails))
            messagebox.showinfo("Success", f"Exported {len(emails)} emails")
    
    def _export_all_emails(self):
        conn = self.db._get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT email FROM customers')
        emails = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        if not emails:
            messagebox.showwarning("Warning", "No customers found")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension='.txt',
            filetypes=[("Text Files", "*.txt"), ("CSV Files", "*.csv")],
            title="Export All Emails"
        )
        
        if file_path:
            with open(file_path, 'w') as f:
                f.write('\n'.join(emails))
            messagebox.showinfo("Success", f"Exported {len(emails)} emails")
    
    def _refresh_customer_count(self):
        self._refresh_customers_list()
    
    def _reset_sent_status(self):
        if messagebox.askyesno("Confirm", "Reset sent status for all customers?"):
            self.db.reset_email_logs()
            self._refresh_customers_list()
            messagebox.showinfo("Success", "Sent status reset for all customers")
    
    def _clear_customers(self):
        if messagebox.askyesno("Confirm", "Delete ALL customers? This cannot be undone!"):
            self.db.clear_customers()
            self._refresh_customers_list()
            messagebox.showinfo("Success", "All customers deleted")
    
    def _upload_stocks_csv(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return
        
        self.attachment_path = file_path
        filename = file_path.split('/')[-1].split('\\')[-1]
        self.attachment_path_label.config(text=filename, foreground='#28a745')
        if hasattr(self, 'stock_file_label'):
            self.stock_file_label.config(text=f"✓ {filename}", foreground='#28a745')
        messagebox.showinfo("Success", f"Stock file selected: {filename}\n\nThis file will be attached to emails.")
    
    def _refresh_stocks_list(self):
        pass
    
    def _clear_stocks(self):
        self.attachment_path = None
        self.attachment_path_label.config(text="No file selected", foreground='#666')
        if hasattr(self, 'stock_file_label'):
            self.stock_file_label.config(text="No file selected", foreground='#666')
    
    def _preview_email(self):
        html = get_stock_email_template_table(None, self.custom_message_entry.get().strip())
        
        self.preview_text.delete('1.0', tk.END)
        self.preview_text.insert('1.0', html[:4000] + "\n\n... (truncated for preview)")
    
    def _start_sending(self):
        accounts = self.db.get_gmail_accounts()
        if not accounts:
            messagebox.showerror("Error", "No Gmail accounts added. Please add accounts first.")
            return
        
        if not self.attachment_path:
            messagebox.showerror("Error", "Please select a stock file to attach.\n\nGo to Stock Data tab and upload your Excel/CSV file.")
            return
        
        if self.send_to_selected_var.get():
            selected_ids = self._get_selected_customer_ids()
            if not selected_ids:
                messagebox.showerror("Error", "No customers selected. Please select customers in the Customers tab.")
                return
            unsent_count = len(self.db.get_unsent_customers_by_ids(selected_ids))
        else:
            unsent_count = self.db.get_unsent_count()
        
        if unsent_count == 0:
            messagebox.showinfo("Info", "All customers have already received emails. Reset to send again.")
            return
        
        self.sending = True
        self.send_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        
        self.selected_customer_ids = self._get_selected_customer_ids() if self.send_to_selected_var.get() else None
        
        threading.Thread(target=self._send_emails_thread, daemon=True).start()
    
    def _send_emails_thread(self):
        try:
            batch_size = int(self.batch_size_entry.get())
            subject = self.subject_entry.get().strip()
            custom_message = self.custom_message_entry.get().strip()
            
            html_template = get_stock_email_template_table(None, custom_message)
            
            total_sent = 0
            total_failed = 0
            
            selected_idx = self.gmail_combo.current()
            if selected_idx < 0 or selected_idx >= len(self.gmail_accounts_list):
                self._update_progress("Please select a Gmail account")
                return
            
            gmail_account = self.gmail_accounts_list[selected_idx]
            
            while self.sending:
                if hasattr(self, 'selected_customer_ids') and self.selected_customer_ids:
                    batch = self.db.get_unsent_customers_by_ids(self.selected_customer_ids)
                    if not batch:
                        batch = self.db.get_customers_by_ids(self.selected_customer_ids)
                        if batch:
                            self._update_progress("All selected customers have received emails.")
                        break
                    batch = batch[:batch_size]
                else:
                    unsent = self.db.get_unsent_count()
                    if unsent == 0:
                        break
                    batch = self.db.get_unsent_customers(limit=batch_size)
                
                if not batch:
                    break
                
                account_id, email, password = gmail_account
                
                sent_today = self.db.get_account_sent_count(account_id)
                if sent_today >= EMAIL_SETTINGS['max_emails_per_account_per_day']:
                    self._update_progress(f"Daily limit reached for {email}")
                    break
                
                self._update_progress(f"Sending batch of {len(batch)} emails via {email}...")
                
                account_data = (account_id, email, password)
                
                for result in self.gmail_service.send_batch(batch, subject, html_template, account_data, attachment_path=self.attachment_path):
                    if not self.sending:
                        break
                    
                    customer_id, status, account_email = result
                    
                    self.db.log_email(customer_id, account_email, status)
                    self.db.increment_email_count(account_id)
                    
                    if status == 'sent':
                        total_sent += 1
                    else:
                        total_failed += 1
                    
                    remaining = self.db.get_unsent_count()
                    self._update_send_stats(total_sent, total_failed, remaining)
                
                if self.sending and self.db.get_unsent_count() > 0:
                    import time
                    delay = int(self.batch_delay_entry.get())
                    self._update_progress(f"Batch complete. Waiting {delay}s before next batch...")
                    time.sleep(delay)
            
            self._update_progress(f"Finished! Sent: {total_sent}, Failed: {total_failed}")
            
        except Exception as e:
            self._update_progress(f"Error: {str(e)}")
        finally:
            self.sending = False
            self.send_btn.config(state='normal')
            self.stop_btn.config(state='disabled')
            self._refresh_accounts_list()
            self._refresh_gmail_dropdown()
    
    def _update_progress(self, message):
        self.root.after(0, lambda: self.progress_label.config(text=message))
    
    def _update_send_stats(self, sent, failed, remaining):
        total = sent + failed + remaining
        if total > 0:
            progress = ((sent + failed) / total) * 100
            self.root.after(0, lambda: self.progress_bar.config(value=progress))
        self.root.after(0, lambda: self.send_stats_label.config(
            text=f"Sent: {sent} | Failed: {failed} | Remaining: {remaining}"
        ))
    
    def _stop_sending(self):
        self.sending = False
        self._update_progress("Stopping...")

def main():
    root = tk.Tk()
    app = StockEmailSenderApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()
